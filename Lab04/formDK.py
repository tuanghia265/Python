from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import re
from openpyxl import *

wb = load_workbook('D:\\-work\Python\\Lab04\\dk.xlsx')

sheet=wb.active

def excel():

    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50
    sheet.column_dimensions['H'].width = 100

    sheet.cell(row=1, column=1).value = "MSSV"
    sheet.cell(row=1, column=2).value = "Họ Ten"
    sheet.cell(row=1, column=3).value = "Ngày sinh"
    sheet.cell(row=1, column=4).value = "Email"
    sheet.cell(row=1, column=5).value = "Số điện thoại"
    sheet.cell(row=1, column=6).value = "Học kỳ"
    sheet.cell(row=1, column=7).value = "Năm học"
    sheet.cell(row=1, column=8).value = "Môn học"

def focus1(event):
    id_entry.focus_set()


def id_check():
    id=id_var.get()

    if len(id)!=7:
        messagebox.showerror("Lỗi","Vui lòng nhập đủ 7 số!!!")
        return 0
    
    for character in id:
        if not character.isdigit():
            messagebox.showerror("Lỗi","Kí tự không hợp lệ")
            return 0
    return 1

def email_check():
    email=email_var.get()

    regex = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"

    if re.match(regex,email):
        return 1
    else:
        messagebox.showerror("Lỗi","Email không hợp lệ!!!")
        return 0

def phone_check():
    phone=phone_var.get()

    if len(phone)!=7:
        messagebox.showerror("Lỗi","Số điện thoại không hợp lệ!!!")
        return 0
    
    for char in phone:
        if not char.isdigit():
            messagebox.showerror("Lỗi","Số điện thoại không hợp lệ!!!")
            return 0
    return 1

def date_check():
    date=date_var.get()

    pattern = r"^\d{2}\/\d{2}\/\d{4}$"
    if re.match(pattern,date):
        return 1
    else:
        messagebox.showerror("Lỗi","Ngày sinh không hợp lệ")
        return 0
    
def hk_check():
    hk=hk_var.get()

    if hk not in['1','2','3']:
        messagebox.showerror("Lỗi","Học kì không hợp lệ!!!")
        hk_var.set("")
    return hk



    
    



if __name__=="__main__":

    root=Tk()
    root.config(bg="light green")
    root.title("Đăng kí học phần")
    root.geometry("800x400")
    Tk=Label(root,text="Thông tin đăng kí học phần",fg="red",bg="light green",font=("Times New Roman",32,'bold')).grid(row=0,column=1)
    id_var=StringVar()
    name_var=StringVar()
    date_var=StringVar()
    email_var=StringVar()
    phone_var=StringVar()
    hk_var=StringVar()
    n=StringVar()
    
    id_label=Label(root,text="Mã số sinh viên:",fg="blue",font=("Times New Roman",12),bg="light green",width=10,anchor='w').grid(column=0,row=1)
    id_entry=Entry(root,textvariable=id_var,font=("Times New Roman",12),width=60).grid(column=1,row=1)

    
    name_label=Label(root,text="Họ tên:",font=("Times New Roman",12),fg="blue",bg="light green",width=10,anchor='w').grid(column=0,row=2)
    name_entry=Entry(root,textvariable=name_var,font=("Times New Roman",12),width=60).grid(column=1,row=2)

    
    date_label=Label(root,text="Ngày sinh:",font=("Times New Roman",12),fg="blue",bg='light green',width=10,anchor='w').grid(column=0,row=3)
    date_entry=Entry(root,textvariable=date_var,font=("Times New Roman",12),width=60).grid(column=1,row=3)

    
    email_label=Label(root,text="Email:",font=("Times New Roman",12),fg='blue',bg='light green',width=10,anchor='w').grid(column=0,row=4)
    email_entry=Entry(root,textvariable=email_var,font=("Times New Roman",12),width=60).grid(column=1,row=4)

    
    phone_label=Label(root,text='Số điện thoại',fg='blue',bg='light green',width=10,anchor='w',font=("Times New Roman",12)).grid(column=0,row=5)
    phone_entry=Entry(root,textvariable=phone_var,font=('Times New Roman',12),width=60).grid(column=1,row=5)

    
    hk_label=Label(root,text='Học kì:',font=('Times New Roman',12),fg='blue',bg='light green',width=10,anchor='w').grid(column=0,row=6)
    hk_entry=Entry(root,textvariable=hk_var,font=('Times New Roman',12),width=60).grid(column=1,row=6)

    
    year_label=Label(root,text='Năm học',fg='blue',bg='light green',font=('Times New Roman',12),width=10,anchor='w').grid(column=0,row=7)
    year_choose=ttk.Combobox(root,font=('Times New Roman',12),width=57,textvariable=n,values=['2021-2022','2022-2023','2023-2024'],state='readonly',background='white').grid(column=1,row=7)
    
    sub_label=Label(root,text='Chọn môn học',fg='blue',bg='light green',font=('Times New Roman',12),width=10,anchor='w').grid(column=0,row=8)

    sub_check_1=Checkbutton(root,text='Lập trình Python',fg='blue',bg='light green',font=('Times New Roman',12),anchor='w').grid(column=1,row=8,padx=0,columnspan=1)
    sub_check_2=Checkbutton(root,text='Lập trình Java',fg='blue',bg='light green',font=('Times New Roman',12),anchor='w').grid(column=2,row=8,padx=0,columnspan=1)
    sub_check_3=Checkbutton(root,text='Công nghệ phần mềm',fg='blue',bg='light green',font=('Times New Roman',12),anchor='w').grid(column=1,row=9,padx=0,columnspan=1)
    sub_check_4=Checkbutton(root,text='Phát triển ứng dụng web',fg='blue',bg='light green',font=('Times New Roman',12),anchor='w').grid(column=2,row=9,padx=0,columnspan=1)

    excel()
    dk_btn=Button(text="Đăng ký",fg='yellow',bg='green',activebackground='red',activeforeground='yellow',cursor='hand2',font=("Times New Roman",12),command=(hk_check)).grid(column=1,row=10)
    dk_btn=Button(text="Thoát",fg='yellow',bg='green',activebackground='red',activeforeground='yellow',cursor='hand2',font=("Times New Roman",12),command=exit).grid(column=2,row=10)

    root.mainloop()